"""
Eksport danych do Excela (.xlsx)
=================================
Tworzy jeden skoroszyt z osobnymi zakładkami: Portfolio, Dziennik,
Historia score. W przeciwieństwie do CSV pozwala na formatowanie
(kolory zysku/straty, nagłówki, auto-szerokość kolumn) i wiele arkuszy
w jednym pliku.

Zwraca bajty pliku (BytesIO), gotowe do st.download_button.
"""
from __future__ import annotations

import io
from datetime import datetime

import pandas as pd

# Kolory (ARGB) do formatowania warunkowego
_GREEN = "FF1A9850"
_RED = "FFD73027"
_HEADER_BG = "FF2563EB"
_HEADER_FG = "FFFFFFFF"


def _style_worksheet(worksheet, df: pd.DataFrame, pnl_columns=None):
    """Stylizuje arkusz: nagłówek, auto-szerokość, kolory dla kolumn P&L."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    header_font = Font(color=_HEADER_FG, bold=True)

    # Nagłówek (wiersz 1)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-szerokość kolumn (na podstawie najdłuższej wartości)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for val in df[col_name]:
            max_len = max(max_len, len(str(val)))
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 3, 50)

    # Kolorowanie wartości P&L (zielony dodatnie, czerwony ujemne)
    pnl_columns = pnl_columns or []
    green_font = Font(color=_GREEN, bold=True)
    red_font = Font(color=_RED, bold=True)
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name not in pnl_columns:
            continue
        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            try:
                if float(cell.value) > 0:
                    cell.font = green_font
                elif float(cell.value) < 0:
                    cell.font = red_font
            except (TypeError, ValueError):
                pass


def build_workbook(
    portfolio_rows: list[dict] | None = None,
    journal_rows: list[dict] | None = None,
    score_history: dict[str, list[dict]] | None = None,
) -> bytes:
    """Buduje skoroszyt .xlsx z dostępnych danych i zwraca bajty.

    portfolio_rows: lista pozycji (z portfolio.analyze_portfolio["positions"])
    journal_rows:   lista wpisów dziennika
    score_history:  {ticker: [{date, score}, ...]} dla zakładki historii
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wrote_any = False

        # --- Portfolio ---
        if portfolio_rows:
            df_pf = pd.DataFrame(portfolio_rows)
            # uporządkuj i nazwij kolumny po polsku, jeśli obecne
            rename = {
                "ticker": "Symbol", "name": "Nazwa", "sector": "Sektor",
                "shares": "Liczba akcji", "buy_price": "Cena zakupu",
                "current_price": "Cena aktualna", "current_value": "Wartość",
                "pnl": "Zysk/strata", "pnl_pct": "Zysk/strata %",
                "score": "Wynik", "currency": "Waluta",
            }
            df_pf = df_pf.rename(columns={k: v for k, v in rename.items() if k in df_pf.columns})
            df_pf.to_excel(writer, sheet_name="Portfolio", index=False)
            _style_worksheet(writer.sheets["Portfolio"], df_pf,
                              pnl_columns=["Zysk/strata", "Zysk/strata %"])
            wrote_any = True

        # --- Dziennik ---
        if journal_rows:
            df_j = pd.DataFrame(journal_rows)
            rename = {
                "entry_date": "Data", "ticker": "Symbol", "decision": "Decyzja",
                "reason": "Powód", "score_at_entry": "Wynik przy wpisie",
                "price_at_entry": "Cena przy wpisie",
            }
            keep = [c for c in rename if c in df_j.columns]
            df_j = df_j[keep].rename(columns=rename)
            df_j.to_excel(writer, sheet_name="Dziennik", index=False)
            _style_worksheet(writer.sheets["Dziennik"], df_j)
            wrote_any = True

        # --- Historia score ---
        if score_history:
            frames = []
            for ticker, rows in score_history.items():
                if not rows:
                    continue
                d = pd.DataFrame(rows)
                d["ticker"] = ticker
                frames.append(d)
            if frames:
                df_h = pd.concat(frames, ignore_index=True)
                df_h = df_h.rename(columns={
                    "date": "Data", "score": "Wynik", "ticker": "Symbol",
                })
                cols = [c for c in ["Data", "Symbol", "Wynik"] if c in df_h.columns]
                df_h = df_h[cols]
                df_h.to_excel(writer, sheet_name="Historia score", index=False)
                _style_worksheet(writer.sheets["Historia score"], df_h)
                wrote_any = True

        # Excel wymaga przynajmniej jednego arkusza
        if not wrote_any:
            pd.DataFrame({"Informacja": ["Brak danych do eksportu"]}).to_excel(
                writer, sheet_name="Info", index=False
            )

    output.seek(0)
    return output.getvalue()


def suggested_filename(prefix: str = "analizator") -> str:
    return f"{prefix}_{datetime.now():%Y%m%d_%H%M}.xlsx"
