"""
Testy dla excel_export.py – budowanie skoroszytu .xlsx.
"""
import io

import openpyxl

from excel_export import build_workbook, suggested_filename


def test_build_workbook_all_sheets():
    data = build_workbook(
        portfolio_rows=[{
            "ticker": "AAPL", "name": "Apple", "sector": "Tech",
            "shares": 5, "buy_price": 150, "current_price": 180,
            "current_value": 900, "pnl": 150, "pnl_pct": 20,
            "score": 65, "currency": "USD",
        }],
        journal_rows=[{
            "entry_date": "2026-01-01", "ticker": "AAPL", "decision": "Kupno",
            "reason": "trend", "score_at_entry": 70, "price_at_entry": 150,
        }],
        score_history={"AAPL": [{"date": "2026-01-01", "score": 60}]},
    )
    assert isinstance(data, bytes)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Portfolio" in wb.sheetnames
    assert "Dziennik" in wb.sheetnames
    assert "Historia score" in wb.sheetnames


def test_build_workbook_empty_has_info_sheet():
    data = build_workbook()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Info" in wb.sheetnames


def test_build_workbook_portfolio_only():
    data = build_workbook(portfolio_rows=[{
        "ticker": "MSFT", "shares": 2, "pnl": -50, "pnl_pct": -10,
    }])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Portfolio" in wb.sheetnames
    assert "Dziennik" not in wb.sheetnames


def test_portfolio_headers_renamed():
    data = build_workbook(portfolio_rows=[{
        "ticker": "AAPL", "shares": 1, "pnl": 10, "pnl_pct": 5,
    }])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Portfolio"]
    headers = [c.value for c in ws[1]]
    assert "Symbol" in headers  # 'ticker' -> 'Symbol'
    assert "Zysk/strata" in headers  # 'pnl' -> 'Zysk/strata'


def test_suggested_filename_format():
    name = suggested_filename("portfolio")
    assert name.startswith("portfolio_")
    assert name.endswith(".xlsx")
